from openpyxl import load_workbook
from pathlib import Path
import json
from datetime import datetime
import openpyxl

def conv(date_or_str: datetime | str, is_full: bool = False) -> str:
    if type(date_or_str) == datetime:
        fmt = r"%d/%m/%Y" if is_full else r"%m/%d" 
        return date_or_str.strftime(fmt)
    else:
        return str(date_or_str)

path = Path(r"C:/Users/ayman/OneDrive/Desktop/crops.xlsx")

wb = load_workbook(filename=path, read_only=True)

sheet = wb["cropFixed"]
rng = sheet["A3":"N176"]

obj = {}

seasons = [cell[0].value for cell in rng]
seasons = list(dict.fromkeys(seasons).keys())
# sheet.cell()

for season in seasons:
    obj[season] = {}

for row in rng:
    row = tuple(r for r in row if r.value != None)
    # print(row[-1].value,len(row))
    # break
    if len(row) <= 10:
        season, crop, traditional_irrigation, modern_irrigation, giving_start, giving_end, due_date, inst_number, inst_date, notes = (c.value for c in row)

        obj[season][crop] = {
            "tr_val": traditional_irrigation, 
            "mod_val": modern_irrigation, 
            "begin" : conv(giving_start), 
            "end": conv(giving_end), 
            "due": conv(due_date), 
            "inst_no": inst_number,
            "inst_dte": conv(inst_date, True),
            "notes": notes
        }
    else:
        season, crop, one, two, three, four, five, six, giving_start, giving_end, due_date, inst_number, inst_date, notes = (c.value for c in row)
        obj[season][crop] = {
            "tr_ind_val": one,
            "mod_ind_val": two,
            "min_tr_comp_val": three,
            "max_tr_comp_val": four,
            "min_mod_comp_val": five,
            "max_mod_comp_val": six,
            "begin" : conv(giving_start), 
            "end": conv(giving_end), 
            "due": conv(due_date), 
            "inst_no": inst_number,
            "inst_dte": conv(inst_date, True),
            "notes": notes
        }


with open("output/out.json", "w") as file:
    json.dump(obj, file, ensure_ascii=False, indent=2)